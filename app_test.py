import os
import cv2
from ultralytics import YOLO
from threading import Thread
import datetime
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl import Workbook
import shutil
from flet import *
import base64

class CamStream:
    def __init__(self, stream_id=0):
        self.stream_id = stream_id
        self.vcap = cv2.VideoCapture(self.stream_id)
        if not self.vcap.isOpened():
            print("[Exiting]: Error accessing stream.")
            exit(0)

        fps_input_stream = int(self.vcap.get(5))
        print(f"FPS of hardware/input stream: {fps_input_stream}")

        self.grabbed, self.frame = self.vcap.read()
        if not self.grabbed:
            print('[Exiting] No more frames to read')
            exit(0)

        self.stopped = True
        self.t = Thread(target=self.update, args=())
        self.t.daemon = True

    def start(self):
        self.stopped = False
        self.t.start()

    def update(self):
        while not self.stopped:
            self.grabbed, self.frame = self.vcap.read()
            if not self.grabbed:
                self.stopped = True
                break
        self.vcap.release()

    def read(self):
        return self.frame

    def stop(self):
        self.stopped = True


class AIDS:
    def __init__(self):
        self.frame_processed = 0
        self.is_target = 'OFF'
        self.model = YOLO('yolov8nbest.pt')
        self.cam_stream = CamStream(0)
        self.cam_stream.start()
        
        self.wb = Workbook()
        self.ws = self.wb.active
        self.image_folder = "temp_images"
        
        if not os.path.exists(self.image_folder):
            os.makedirs(self.image_folder)
        
        self.is_running = False
        self.is_loading = False

    def predict(self, img, conf=0.5):
        results = self.model(img, conf=conf, verbose=False)
        if not results or len(results) == 0:
            return img

        current_time = datetime.datetime.now().strftime("%I:%M%p")
        for result in results:
            count = result.boxes.shape[0]
            for i in range(count):
                cls = int(result.boxes.cls[i].item())
                name = result.names[cls]
                confidence = float(result.boxes.conf[i].item())
                bbox = result.boxes.xywh[i].cpu().numpy()
                x, y, w, h = bbox
                cv2.rectangle(img, (int(x - w/2), int(y - h/2)), (int(x + w/2), int(y + h/2)), (255, 0, 255), 1)
                cv2.putText(img, name + ':' + str(round(confidence, 2)), (int(bbox[0]), int(bbox[1] - 40)),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 255), 2)
                self.write_to_excel(name, img, current_time)
                self.is_target_object(name)

        return img

    def is_target_object(self, name):
        if name in ['nohandrailmidleft', 'nohandrailleftfar', 'nohandrailmidright', 'nohandrailrightfar', 'nohandrailupleft', 'nohandraillowright']:
            self.is_target = '1'

    def write_to_excel(self, data, img, current_time):
        img_filename = f"{self.image_folder}/frame_image_{self.frame_processed}.png"
        cv2.imwrite(img_filename, img)
        img = ExcelImage(img_filename)

        self.ws.append([data, current_time])
        self.ws.add_image(img, 'C' + str(self.ws.max_row))

        self.adjust_dimensions()

    def adjust_dimensions(self):
        for col in self.ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            column = col[0].column_letter
            self.ws.column_dimensions[column].width = max_length + 2

        for row in self.ws.iter_rows():
            for cell in row:
                self.ws.row_dimensions[cell.row].height = 300

    def export_data(self):
        self.wb.save('handrail.xlsx')
        shutil.rmtree(self.image_folder)
        print("Data exported successfully")

    def save_frame(self, frame):
        _, im_arr = cv2.imencode('.png', frame)
        im_b64 = base64.b64encode(im_arr).decode('utf-8')
        return im_b64

    def start_detection(self, video):
        while True:
            if self.cam_stream.stopped:
                break
            frame = self.cam_stream.read()
            frame = cv2.resize(frame, (416, 416))
            frame = self.predict(frame)
            self.frame_processed += 1

            frame_b64 = self.save_frame(frame)
            video.src = f"data:image/png;base64,{frame_b64}"
            video.update()

        self.cam_stream.stop()
        cv2.destroyAllWindows()
    
    def start_or_stop_app(self, e, start_stop_button_ref, video):
        if self.is_running:
            self.page.dialog = self.page.stopped_app
            self.page.dialog.open = True
            start_stop_button_ref.current.text = "Start"
            start_stop_button_ref.current.bgcolor = colors.GREEN
        else:
            self.is_loading = True
            self.page.dialog = self.page.started_app
            self.page.dialog.open = True
            start_stop_button_ref.current.text = "Stop"
            start_stop_button_ref.current.bgcolor = colors.RED

            self.detection_app.start_detection(video)

            self.is_loading = False

        start_stop_button_ref.current.update()
        self.page.update()
        self.is_running = not self.is_running
    
    def build(self):
        self.img = Image(
            width=400,
            height=300,
            border_radius=border_radius.all(16),
            color=colors.BLACK,
            fit=ImageFit.CONTAIN
        )

        return self.img


def main(page: Page):
    app = AIDS()
    page.title = 'AIDS'
    page.vertical_alignment = MainAxisAlignment.CENTER
    page.bgcolor = colors.WHITE

    video = app.build()

    start_stop_button_ref = Ref[FilledButton]()

    indicator = Container(
        width=25,
        height=25,
        bgcolor=colors.RED,
        border_radius=25,
    )

    username = TextField(label='Username', border=InputBorder.NONE, filled=True, prefix_icon=icons.PERSON)
    password = TextField(label='Password', password=True, can_reveal_password=True, border=InputBorder.NONE, filled=True, prefix_icon=icons.LOCK)

    loading_spinner = Row(
        [
            ProgressRing(stroke_width=2, color=colors.BLUE_500),
            Text('Loading...')
        ],
        visible=False
    )

    page.connect_success = AlertDialog(title=Text("Device is connected"))
    page.connect_failed = AlertDialog(title=Text("Device connection failed. Please try again"))
    page.login_success = AlertDialog(title=Text("Logged In"))
    page.login_failed = AlertDialog(title=Text("Login failed. Please try again"))
    page.started_app = AlertDialog(title=Text("AIDS Starting"))
    page.stopped_app = AlertDialog(title=Text("AIDS Stopped"))
    page.exported_data = AlertDialog(title=Text("Data Exported"))

    def login(e):
        if username.value == 'admin' and password.value == '123poleng':
            page.dialog = page.login_success
            page.dialog.open = True
            start_stop_button_ref.current.disabled = False
        else:
            page.dialog = page.login_failed
            page.dialog.open = True

        page.update()

    def feedback_test(e):
        page.dialog = page.connect_success
        page.dialog.open = True
        indicator.bgcolor = colors.GREEN  # Change indicator color to green
        indicator.update()
        page.update()

    page.add(
        Container(
            content=Row(
                [
                    Container(
                        width=400,
                        content=Column(
                            [
                                Image(src='pertamina.png', width=100),
                                Text('PERTAMINA \n"AIDS"', style=TextStyle(weight=FontWeight.W_800, color=colors.BLACK, size=50)),
                                username,
                                password,
                                FilledButton(
                                    'Login',
                                    on_click=login,
                                    adaptive=True,
                                    width=500,
                                    style=ButtonStyle(
                                        color=colors.WHITE,
                                        bgcolor=colors.RED,
                                        padding=padding.symmetric(vertical=20)
                                    )
                                ),
                                FilledButton(
                                    'Start',
                                    on_click=lambda e: app.start_or_stop_app(e, start_stop_button_ref, page, video),
                                    adaptive=True,
                                    ref=start_stop_button_ref,
                                    width=500,
                                    disabled=True,
                                    style=ButtonStyle(
                                        color=colors.WHITE,
                                        bgcolor=colors.GREEN,
                                        padding=padding.symmetric(vertical=20)
                                    )
                                ),
                                OutlinedButton(
                                    'Export Data',
                                    on_click=app.export_data,
                                    adaptive=True,
                                    width=500,
                                    style=ButtonStyle(
                                        color=colors.BLACK,
                                        padding=padding.symmetric(vertical=20)
                                    )
                                ),
                            ],
                            spacing=16
                        ),
                    ),
                    # Right side: Video frame and feedback
                    Column(
                        [
                            AIDS(),
                            Row(
                                [
                                    OutlinedButton(
                                        'Feedback Test',
                                        on_long_press=feedback_test,
                                        adaptive=True,
                                        width=150,
                                        style=ButtonStyle(
                                            color=colors.BLACK,
                                            padding=padding.symmetric(vertical=20)
                                        )
                                    ),
                                    indicator  # Light indicator container
                                ],
                                alignment=MainAxisAlignment.SPACE_BETWEEN,
                                width=400
                            ),
                            loading_spinner  # Show loading spinner if is_loading=True
                        ],
                        spacing=16
                    ),
                ],
                alignment=MainAxisAlignment.CENTER,
                vertical_alignment=CrossAxisAlignment.CENTER,
                spacing=50,
            ),
            padding=padding.all(10),
            alignment=alignment.center
        )
    )

if __name__ == "__main__":
    app(target=main)
