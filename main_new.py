import os
import cv2
import numpy as np
from ultralytics import YOLO
from threading import Thread
import time
import datetime
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import shutil
import serial

frame_processed = 0
is_target = 'OFF'

# CamStream class for video stream handling
class CamStream:
    def __init__(self, stream_id=0):
        self.stream_id = stream_id
        self.vcap = cv2.VideoCapture(self.stream_id)
        if not self.vcap.isOpened():
            print("[Exiting]: Error accessing stream.")
            exit(0)

        fps_input_stream = int(self.vcap.get(5))
        print(f"FPS of hardware/input stream: {fps_input_stream}")

        self.grabbed, self.frame = self.vcap.read()
        if not self.grabbed:
            print('[Exiting] No more frames to read')
            exit(0)

        self.stopped = True
        self.t = Thread(target=self.update, args=())
        self.t.daemon = True

    def start(self):
        self.stopped = False
        self.t.start()

    def update(self):
        while not self.stopped:
            self.grabbed, self.frame = self.vcap.read()
            if not self.grabbed:
                self.stopped = True
                break
        self.vcap.release()

    def read(self):
        return self.frame

    def stop(self):
        self.stopped = True

def predict(model, img, frame_count, conf=0.5):
    global ref_image
    global ser
    crop_img  = img.copy()
    results = model(img, conf=conf, verbose=False)
    if not results or len(results) == 0:
        return img

    current_time = datetime.datetime.now().strftime("%I:%M%p")
    for result in results:
        count = result.boxes.shape[0]
        for i in range(count):
            cls = int(result.boxes.cls[i].item())
            name = result.names[cls]
            if name == 'person':
                confidence = float(result.boxes.conf[i].item())
                bbox = result.boxes.xywh[i].cpu().numpy()
                x, y, w, h = bbox
                cv2.rectangle(img, (int(x - w/2), int(y - h/2)), (int(x + w/2), int(y + h/2)), (255, 0, 255), 1)
                cv2.putText(img, name + ':' + str(round(confidence, 2)), (int(bbox[0]), int(bbox[1] - 40)),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 255), 2)
                pts1 = np.array([[255, 234], [260, 234], [307, 130], [302, 131]])
                pts2 = np.array([[311, 113], [404, 91], [404, 95], [311, 117]])
                crop_img = crop(crop_img, pts1, pts2)
                cv2.imwrite('test1.png', crop_img[int(y - h/2):int(y + h/2),int(x - w/2):int(x + w/2)])
                cv2.imwrite('test2.png', ref_image[int(y - h/2):int(y + h/2),int(x - w/2):int(x + w/2)])
                if calculate_pixel(crop_img[int(y - h/2):int(y + h/2), int(x - w/2):int(x + w/2)]) > calculate_pixel(ref_image[int(y - h/2):int(y + h/2), int(x - w/2):int(x + w/2)]):
                    ser.write(('on\n').encode('utf-8'))  # Send data
                write_to_excel(name, img, current_time, frame_count)

    return img

def crop(frame, pts1, pts2):

    # Create a mask of the same size as the image, initialized with zeros (black)
    mask = np.zeros(frame.shape[:2], dtype=np.uint8)

   # Fill the two polygons on the mask with white (255)
    cv2.fillPoly(mask, [pts1], 255)
    cv2.fillPoly(mask, [pts2], 255)

    # Apply the mask to the image
    masked_image = cv2.bitwise_and(frame, frame, mask=mask)

    return masked_image

def calculate_pixel(frame):
    return np.std(frame)

def write_to_excel(data, img, current_time, frame_count):
    img_filename = f"{image_folder}/frame_image_{frame_count}.png"
    cv2.imwrite(img_filename, img)
    img = Image(img_filename)

    ws.append([data, current_time])
    ws.add_image(img, 'C' + str(ws.max_row))

    adjust_dimensions(ws)

def adjust_dimensions(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_length + 2

    for row in ws.iter_rows():
        for cell in row:
            ws.row_dimensions[cell.row].height = 300

def export_data():
    wb.save('handrail.xlsx')
    shutil.rmtree(image_folder)
    print("Data exported successfully")

def save_frame(frame, frame_count):
    img_filename = f"frames/latest_frame.png"
    cv2.imwrite(img_filename, frame)

def main(model):
    global frame_processed
    while True:
        if cam_stream.stopped:
            break
        frame = cam_stream.read()
        frame = cv2.resize(frame, (416, 416))
        frame = predict(model, frame, frame_processed)
        frame_processed += 1
        save_frame(frame, frame_processed)
        cv2.imshow('cctv', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    cam_stream.stop()
    cv2.destroyAllWindows()


# Initialization
model = YOLO('yolov8n.pt')
cam_stream = CamStream('test.mp4')
cam_stream.start()

wb = Workbook()
ws = wb.active

image_folder = "temp_images"
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

# Set up the serial connection
ser = serial.Serial('COM6', 115200, timeout=1)

ref_image = cv2.imread('clean_handrail.jpg')

if __name__ == "__main__":
    main(model)
