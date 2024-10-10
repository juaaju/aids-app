import os
import cv2
from ultralytics import YOLO
from threading import Thread
import datetime
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import shutil
from flet import *
import base64
import serial
import numpy as np
from bleak import BleakClient
import asyncio
import time

# BLE device and characteristic details
DEVICE_ADDRESS = "A0:A3:B3:2A:D8:22"  # MAC address of your ESP32
SERVICE_UUID = "4fafc201-1fb5-459e-8fcc-c5c9c331914b"
CHARACTERISTIC_UUID = "beb5483e-36e1-4688-b7f5-ea07361b26a8"

client = None

# CamStream class for video stream handling
class CamStream:
    def __init__(self, stream_id=0):
        self.stream_id = stream_id
        self.vcap = cv2.VideoCapture(self.stream_id)
        if not self.vcap.isOpened():
            print("[Exiting]: Error accessing stream.")
            exit(0)

        fps_input_stream = int(self.vcap.get(5))
        print(f"FPS of hardware/input stream: {fps_input_stream}")

        self.grabbed, self.frame = self.vcap.read()
        if not self.grabbed:
            print('[Exiting] No more frames to read')
            exit(0)

        self.stopped = True
        self.t = Thread(target=self.update, args=())
        self.t.daemon = True

    def start(self):
        self.stopped = False
        self.t.start()

    def update(self):
        while not self.stopped:
            self.grabbed, self.frame = self.vcap.read()
            if not self.grabbed:
                self.stopped = True
                break
        self.vcap.release()

    def read(self):
        return self.frame

    def stop(self):
        self.stopped = True

async def setup_ble_client():
    global client
    try:
        client = BleakClient(DEVICE_ADDRESS)
        await client.connect()
        if client.is_connected:
            print("Connected to BLE device")
        else:
            print("Failed to connect to BLE device")
    except Exception as e:
        print(f"Error connecting to BLE device: {e}")

async def send_ble_signal():
    global client
    if client and client.is_connected:
        try:
            await client.write_gatt_char(CHARACTERISTIC_UUID, b'on\n')  # Send 'on' signal
            print("Signal sent over BLE")
        except Exception as e:
            print(f"Error sending BLE signal: {e}")
    else:
        print("BLE client is not connected")

async def predict(model, img, frame_count, conf=0.5):
    global ref_image
    crop_img = img.copy()
    results = model(img, conf=conf, verbose=False)
    if not results or len(results) == 0:
        return img

    current_time = datetime.datetime.now().strftime("%I:%M%p")

    is_send = False
    for result in results:
        count = result.boxes.shape[0]
        for i in range(count):
            cls = int(result.boxes.cls[i].item())
            name = result.names[cls]
            if name == 'person':
                confidence = float(result.boxes.conf[i].item())
                bbox = result.boxes.xywh[i].cpu().numpy()
                x, y, w, h = bbox
                lx = int(x - w / 2)
                ux = int(x + w / 2)
                ly = int(y - h / 2)
                uy = int(y + h / 2)
                cv2.rectangle(img, (lx, ly), (ux, uy), (255, 0, 255), 1)
                cv2.putText(img, name + ':' + str(round(confidence, 2)), (int(bbox[0]), int(bbox[1] - 40)),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 255), 2)

                # Adjust this condition based on your cropping logic
                if lx >= 200 and ly >= 50 and uy <= 300:
                    pts1 = np.array([[256, 234], [258, 234], [305, 132], [303, 132]])
                    pts2 = np.array([[312, 113], [404, 92], [404, 94], [312, 115]])
                    crop_img = crop(crop_img, pts1, pts2)

                    # Check if the person is holding the handrail
                    if calculate_pixel(crop_img[ly:uy, lx:ux]) <= calculate_pixel(ref_image[ly:uy, lx:ux]):
                        is_send = True
        if is_send:
            await send_ble_signal()  # Await the asynchronous function
            write_to_excel(name, img, current_time, frame_count)

    return img

def crop(frame, pts1, pts2):

    # Create a mask of the same size as the image, initialized with zeros (black)
    mask = np.zeros(frame.shape[:2], dtype=np.uint8)

   # Fill the two polygons on the mask with white (255)
    cv2.fillPoly(mask, [pts1], 255)
    cv2.fillPoly(mask, [pts2], 255)

    # Apply the mask to the image
    masked_image = cv2.bitwise_and(frame, frame, mask=mask)

    return masked_image

def calculate_pixel(frame):
    return np.std(frame)

def write_to_excel(data, img, current_time, frame_count):
    img_filename = f"{image_folder}/frame_image_{frame_count}.png"
    cv2.imwrite(img_filename, img)
    img = Image(img_filename)

    ws.append([data, current_time])
    ws.add_image(img, 'C' + str(ws.max_row))

    adjust_dimensions(ws)

def adjust_dimensions(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_length + 2

    for row in ws.iter_rows():
        for cell in row:
            ws.row_dimensions[cell.row].height = 300

def export_data(e):
    global frame_processed  # Track the number of processed frames
    if frame_processed > 0:  # Check if there are frames to export
        wb.save('handrail.xlsx')  # Save the workbook
        shutil.rmtree(image_folder)  # Clean up temporary images


def save_frame(frame):
    _, im_arr = cv2.imencode('.jpg', frame)
    im_b64 = base64.b64encode(im_arr)
    return im_b64.decode('utf-8')

async def start_detection(model, video):
    global frame_b64
    global frame_processed
    while True:
        if cam_stream.stopped:
            break
        frame = cam_stream.read()
        frame = cv2.resize(frame, (416, 416))
        frame = await predict(model, frame, frame_processed)
        frame_processed += 1

        if frame_processed == 1:
            video.src = None
        video.src_base64 = save_frame(frame)
        video.update()
    
    cam_stream.stop()
    cv2.destroyAllWindows()

def main(page: Page):
    page.title = 'AIDS'
    page.vertical_alignment = MainAxisAlignment.CENTER
    page.bgcolor = colors.WHITE

    # Track button states
    is_running = False  # Track if the app is running
    is_loading = False  # Track if loading spinner should be shown
    start_stop_button_ref = Ref[FilledButton]()

    # Light indicator container
    indicator = Container(
        width=25,
        height=25,
        bgcolor=colors.RED,  # Start with red
        border_radius=25,
    )

    # Username and password fields
    username = TextField(label='Username', border=InputBorder.NONE, filled=True, prefix_icon=icons.PERSON)
    password = TextField(label='Password', password=True, can_reveal_password=True, border=InputBorder.NONE, filled=True, prefix_icon=icons.LOCK)

    # Loading spinner (only shown when is_loading=True)
    loading_spinner = Row(
        [
            ProgressRing(stroke_width=2, color=colors.BLUE_500),
            Text('Loading...')
        ],
        visible=False
    )

    result_video = Image(
        width=400,
        height=300,
        border_radius=border_radius.all(16),
        fit=ImageFit.CONTAIN,
        src_base64=None,
        src='sample_image.png'
    )

    # Dialogs
    connect_success = AlertDialog(title=Text("Device is connected"))
    login_success = AlertDialog(title=Text("Logged In"))
    login_failed = AlertDialog(title=Text("Login failed. Please try again"))

    def start_detection_thread(model, video):
        global detection_thread
        # Start detection in a new thread to prevent blocking
        detection_thread = Thread(target=start_detection, args=(model, video))
        detection_thread.start()

    # Event Handlers
    def start_or_stop_app(e):
        global detection_thread
        nonlocal is_running
        nonlocal is_loading
        if is_running:
            # Change to "Start" state
            cam_stream.stop()
            start_stop_button_ref.current.text = "Start"
            start_stop_button_ref.current.style.bgcolor = colors.GREEN

            if detection_thread is not None:
                detection_thread.join()  # Ensure thread completes before moving on
                detection_thread = None
        else:
            # Start main.py and video updates
            start_stop_button_ref.current.text = "Stop"
            start_stop_button_ref.current.style.bgcolor = colors.RED
            # is_loading = True
            # loading_spinner.visible = True
            page.update()

            detection_thread = Thread(target=asyncio.run, args=(start_detection(model, result_video),))
            detection_thread.start()

            # is_loading = False
            # loading_spinner.visible = False

        start_stop_button_ref.current.update()
        page.update()
        is_running = not is_running

    async def feedback_test(e):
        await send_ble_signal()
        page.open(connect_success)
        indicator.bgcolor = colors.GREEN  # Change indicator color to green
        indicator.update()
        page.update()

    def login(e):
        if username.value == 'admin' and password.value == '123poleng':
            page.open(login_success)
            start_stop_button_ref.current.disabled = False  # Enable the Start/Stop button
            start_stop_button_ref.current.update()
        else:
            page.open(login_failed)
            password.value = ""
            password.update()

        page.update()

    # Layout setup
    page.add(
        Container(
            content=Row(
                [
                    # Left side: Login section
                    Container(
                        width=400,
                        content=Column(
                            [
                                Image(src='pertamina.png', width=100),
                                Text('PERTAMINA \n"AIDS"', style=TextStyle(weight=FontWeight.W_800, color=colors.BLACK, size=50)),
                                username,
                                password,
                                FilledButton(
                                    'Login',
                                    on_click=login,
                                    adaptive=True,
                                    width=500,
                                    style=ButtonStyle(
                                        color=colors.WHITE,
                                        bgcolor=colors.RED,
                                        padding=padding.symmetric(vertical=20)
                                    )
                                ),
                                FilledButton(
                                    'Start',
                                    on_click=start_or_stop_app,
                                    adaptive=True,
                                    ref=start_stop_button_ref,
                                    width=500,
                                    disabled=True,  # Initially disabled
                                    style=ButtonStyle(
                                        color=colors.WHITE,
                                        bgcolor=colors.GREEN,
                                        padding=padding.symmetric(vertical=20)
                                    )
                                ),
                                OutlinedButton(
                                    'Export Data',
                                    on_click=export_data,
                                    adaptive=True,
                                    width=500,
                                    style=ButtonStyle(
                                        color=colors.BLACK,
                                        padding=padding.symmetric(vertical=20)
                                    )
                                ),
                            ],
                            spacing=16
                        ),
                    ),
                    # Right side: Video frame and feedback
                    Column(
                        [
                            Container(
                                alignment=alignment.center,
                                border_radius=border_radius.all(20),
                                bgcolor=colors.BLACK,
                                width=400,
                                height=300,
                                content=result_video
                            ),
                            Row(
                                [
                                    OutlinedButton(
                                        'Feedback Test',
                                        on_long_press=feedback_test,
                                        adaptive=True,
                                        width=150,
                                        style=ButtonStyle(
                                            color=colors.BLACK,
                                            padding=padding.symmetric(vertical=20)
                                        )
                                    ),
                                    indicator  # Light indicator container
                                ],
                                alignment=MainAxisAlignment.SPACE_BETWEEN,
                                width=400
                            ),
                            loading_spinner  # Show loading spinner if is_loading=True
                        ],
                        spacing=16
                    ),
                ],
                alignment=MainAxisAlignment.CENTER,
                vertical_alignment=CrossAxisAlignment.CENTER,
                spacing=50,
            ),
            padding=padding.all(10),
            alignment=alignment.center
        )
    )

frame_processed = 0
video_path = "rtsp://admin:pertamina321@10.205.64.111:554/Streaming/Channels/301"

model = YOLO('yolov8n.pt')
cam_stream = CamStream('test.mp4')
cam_stream.start()

wb = Workbook()
ws = wb.active

image_folder = "temp_images"
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

# Set up the serial connection
# ser = serial.Serial('COM5', 115200, timeout=1)

ref_image = cv2.imread('clean_handrail.png')

if __name__ == "__main__":
    # Run the BLE setup in an asyncio loop
    loop = asyncio.get_event_loop()
    loop.run_until_complete(setup_ble_client())

    app(main)
    shutil.rmtree(image_folder)
