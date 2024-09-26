import os
import cv2
import numpy as np
from ultralytics import YOLO
from threading import Thread
import time
from gtts import gTTS
from playsound import playsound
import datetime
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import asyncio
from bleak import BleakClient
import shutil
import ble_utils

# BLE device and characteristic details
ESP32_ADDRESS = "A0:A3:B3:2A:D8:22"  # MAC address of your ESP32
SERVICE_UUID = "4fafc201-1fb5-459e-8fcc-c5c9c331914b"
CHARACTERISTIC_UUID = "beb5483e-36e1-4688-b7f5-ea07361b26a8"

is_target = 'OFF'

# CamStream class for video stream handling
class CamStream:
    def __init__(self, stream_id=0):
        self.stream_id = stream_id
        self.vcap = cv2.VideoCapture(self.stream_id)
        if not self.vcap.isOpened():
            print("[Exiting]: Error accessing stream.")
            exit(0)

        fps_input_stream = int(self.vcap.get(5))
        print(f"FPS of hardware/input stream: {fps_input_stream}")

        self.grabbed, self.frame = self.vcap.read()
        if not self.grabbed:
            print('[Exiting] No more frames to read')
            exit(0)

        self.stopped = True
        self.t = Thread(target=self.update, args=())
        self.t.daemon = True

    def start(self):
        self.stopped = False
        self.t.start()

    def update(self):
        while not self.stopped:
            self.grabbed, self.frame = self.vcap.read()
            if not self.grabbed:
                self.stopped = True
                break
        self.vcap.release()

    def read(self):
        return self.frame

    def stop(self):
        self.stopped = True

async def predict(model, img, frame_count, client, conf=0.5):
    results = model(img, conf=conf, verbose=False)
    if not results or len(results) == 0:
        return img

    current_time = datetime.datetime.now().strftime("%I:%M%p")
    for result in results:
        count = result.boxes.shape[0]
        for i in range(count):
            cls = int(result.boxes.cls[i].item())
            name = result.names[cls]
            confidence = float(result.boxes.conf[i].item())
            bbox = result.boxes.xywh[i].cpu().numpy()
            x, y, w, h = bbox
            cv2.rectangle(img, (int(x - w/2), int(y - h/2)), (int(x + w/2), int(y + h/2)), (255, 0, 255), 1)
            cv2.putText(img, name + ':' + str(round(confidence, 2)), (int(bbox[0]), int(bbox[1] - 40)),
                        cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 255), 2)
            write_to_excel(name, img, current_time, frame_count)
            await is_target_object(name, client)

    return img

async def is_target_object(name, client):
    global is_target
    if name in ['nohandrailmidleft', 'nohandrailleftfar', 'nohandrailmidright', 'nohandrailrightfar', 'nohandrailupleft', 'nohandraillowright']:
        is_target = '1'
    await ble_send(CHARACTERISTIC_UUID, is_target, client)

async def ble_send(characteristic_uuid, command, ble_client=None):
    await ble_client.write_gatt_char(characteristic_uuid, command.encode())

def write_to_excel(data, img, current_time, frame_count):
    img_filename = f"{image_folder}/frame_image_{frame_count}.png"
    cv2.imwrite(img_filename, img)
    img = Image(img_filename)

    ws.append([data, current_time])
    ws.add_image(img, 'C' + str(ws.max_row))

    adjust_dimensions(ws)

def adjust_dimensions(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_length + 2

    for row in ws.iter_rows():
        for cell in row:
            ws.row_dimensions[cell.row].height = 300

def sound_notification(text):
    tts = gTTS(text, lang='id', slow=False)
    tts.save('speech.mp3')
    playsound('speech.mp3')

async def main(model, frame_processed):
    async with BleakClient(ESP32_ADDRESS) as client:
        try:
            while True:
                if cam_stream.stopped:
                    break
                frame = cam_stream.read()
                frame = cv2.resize(frame, (416, 416))
                frame = predict(model, frame, frame_processed, client)
                frame_processed += 1
                cv2.imshow('cctv', frame)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        finally:
            cam_stream.stop()
            cv2.destroyAllWindows()
            wb.save('handrail.xlsx')
            shutil.rmtree(image_folder)


# Initialization
model = YOLO('yolov8nbest.pt')
cam_stream = CamStream("00000000815000000.mp4")
cam_stream.start()

wb = Workbook()
ws = wb.active

image_folder = "temp_images"
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

frame_processed = 0

if __name__ == "__main__":
    asyncio.run(main(model, frame_processed))

# ble_client = asyncio.run(ble_utils.connect(ESP32_ADDRESS))
