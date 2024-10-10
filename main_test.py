import os
import cv2
import numpy as np
from ultralytics import YOLO
from threading import Thread
import time
import datetime
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import shutil
import serial
from bleak import BleakClient
import asyncio

# BLE device and characteristic details
DEVICE_ADDRESS = "A0:A3:B3:2A:D8:22"  # MAC address of your ESP32
SERVICE_UUID = "4fafc201-1fb5-459e-8fcc-c5c9c331914b"
CHARACTERISTIC_UUID = "beb5483e-36e1-4688-b7f5-ea07361b26a8"

client = None

frame_processed = 0
is_target = 'OFF'

# CamStream class for video stream handling
class CamStream:
    def __init__(self, stream_id=0):
        self.stream_id = stream_id
        self.vcap = cv2.VideoCapture(self.stream_id)
        if not self.vcap.isOpened():
            print("[Exiting]: Error accessing stream.")
            exit(0)

        fps_input_stream = int(self.vcap.get(5))
        print(f"FPS of hardware/input stream: {fps_input_stream}")

        self.grabbed, self.frame = self.vcap.read()
        if not self.grabbed:
            print('[Exiting] No more frames to read')
            exit(0)

        self.stopped = True
        self.t = Thread(target=self.update, args=())
        self.t.daemon = True

    def start(self):
        self.stopped = False
        self.t.start()

    def update(self):
        while not self.stopped:
            self.grabbed, self.frame = self.vcap.read()
            if not self.grabbed:
                self.stopped = True
                break
        self.vcap.release()

    def read(self):
        return self.frame

    def stop(self):
        self.stopped = True

async def setup_ble_client():
    global client
    try:
        client = BleakClient(DEVICE_ADDRESS)
        await client.connect()
        if client.is_connected:
            print("Connected to BLE device")
        else:
            print("Failed to connect to BLE device")
    except Exception as e:
        print(f"Error connecting to BLE device: {e}")

async def send_ble_signal():
    global client
    if client and client.is_connected:
        try:
            await client.write_gatt_char(CHARACTERISTIC_UUID, b'on\n')  # Send 'on' signal
            print("Signal sent over BLE")
        except Exception as e:
            print(f"Error sending BLE signal: {e}")
    else:
        print("BLE client is not connected")

async def predict(model, img, frame_count, conf=0.5):
    global ref_image
    crop_img = img.copy()
    results = model(img, conf=conf, verbose=False)
    if not results or len(results) == 0:
        return img

    current_time = datetime.datetime.now().strftime("%I:%M%p")
    for result in results:
        count = result.boxes.shape[0]
        for i in range(count):
            cls = int(result.boxes.cls[i].item())
            name = result.names[cls]
            if name == 'person':
                confidence = float(result.boxes.conf[i].item())
                bbox = result.boxes.xywh[i].cpu().numpy()
                x, y, w, h = bbox
                lx = int(x - w / 2)
                ux = int(x + w / 2)
                ly = int(y - h / 2)
                uy = int(y + h / 2)
                cv2.rectangle(img, (lx, ly), (ux, uy), (255, 0, 255), 1)
                cv2.putText(img, name + ':' + str(round(confidence, 2)), (int(bbox[0]), int(bbox[1] - 40)),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 255), 2)

                # Adjust this condition based on your cropping logic
                if lx >= 200 and ly >= 50 and uy <= 300:
                    pts1 = np.array([[256, 234], [258, 234], [305, 132], [303, 132]])
                    pts2 = np.array([[312, 113], [404, 92], [404, 94], [312, 115]])
                    crop_img = crop(crop_img, pts1, pts2)

                    # Check if the person is holding the handrail
                    if calculate_pixel(crop_img[ly:uy, lx:ux]) <= calculate_pixel(ref_image[ly:uy, lx:ux]):
                        await send_ble_signal()  # Await the asynchronous function
                    write_to_excel(name, img, current_time, frame_count)

    return img

def crop(frame, pts1, pts2):

    # Create a mask of the same size as the image, initialized with zeros (black)
    mask = np.zeros(frame.shape[:2], dtype=np.uint8)

   # Fill the two polygons on the mask with white (255)
    cv2.fillPoly(mask, [pts1], 255)
    cv2.fillPoly(mask, [pts2], 255)

    # Apply the mask to the image
    masked_image = cv2.bitwise_and(frame, frame, mask=mask)

    return masked_image

def calculate_pixel(frame):
    return np.std(frame)

def write_to_excel(data, img, current_time, frame_count):
    img_filename = f"{image_folder}/frame_image_{frame_count}.png"
    cv2.imwrite(img_filename, img)
    img = Image(img_filename)

    ws.append([data, current_time])
    ws.add_image(img, 'C' + str(ws.max_row))

    adjust_dimensions(ws)

def adjust_dimensions(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_length + 2

    for row in ws.iter_rows():
        for cell in row:
            ws.row_dimensions[cell.row].height = 300

def export_data():
    wb.save('handrail.xlsx')
    shutil.rmtree(image_folder)
    print("Data exported successfully")

def save_frame(frame, frame_count):
    img_filename = f"frames/latest_frame.png"
    cv2.imwrite(img_filename, frame)

async def main(model):
    global frame_processed
    while True:
        if cam_stream.stopped:
            break
        frame = cam_stream.read()
        frame = cv2.resize(frame, (416, 416))
        frame = await predict(model, frame, frame_processed)
        frame_processed += 1
        save_frame(frame, frame_processed)
        cv2.imshow('cctv', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    cam_stream.stop()
    cv2.destroyAllWindows()


# Initialization
model = YOLO('yolov8n.pt')
cam_stream = CamStream('test.mp4')
cam_stream.start()

wb = Workbook()
ws = wb.active

image_folder = "temp_images"
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

# Set up the serial connection
# ser = serial.Serial('COM5', 115200, timeout=1)

ref_image = cv2.imread('clean_handrail.png')

if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.run_until_complete(setup_ble_client())
    asyncio.run(main(model))